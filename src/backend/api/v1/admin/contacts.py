"""Admin Authorized Contacts endpoints (Sprint 18).

Directory of authorized users with extended fields (name, organization, country,
sector, registered_by) and per-frontend overrides (replace / append).
Supports Excel/CSV import/export with additive merge.
"""

import csv
import io
import json
import logging
from typing import Any, Literal

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from src.api.v1.admin.auth import require_admin
from src.services.smtp_service import (
    load_authorized_contacts,
    save_authorized_contacts,
    _CONTACT_FIELDS,
    _OVERRIDE_MODES,
    _normalise_contact,
)

logger = logging.getLogger("backend.admin.contacts")

router = APIRouter(prefix="/admin/contacts", tags=["admin-contacts"])


# --- Schemas ---

class Contact(BaseModel):
    email: str
    first_name: str = ""
    last_name: str = ""
    organization: str = ""
    country: str = ""
    sector: str = ""
    registered_by: str = ""


class GlobalContactsRequest(BaseModel):
    contacts: list[Contact]


class FrontendOverrideRequest(BaseModel):
    mode: Literal["replace", "append"] = "replace"
    contacts: list[Contact] = Field(default_factory=list)


# --- Read ---

@router.get("")
async def get_contacts(_: dict = Depends(require_admin)):
    """Return full contacts store (global + per-frontend overrides)."""
    return load_authorized_contacts()


# --- Global writes ---

@router.put("/global")
async def update_global_contacts(req: GlobalContactsRequest, _: dict = Depends(require_admin)):
    store = load_authorized_contacts()
    store["global"] = [c.model_dump() for c in req.contacts]
    clean = save_authorized_contacts(store)
    return {"global": clean["global"]}


# --- Per-frontend overrides ---

@router.put("/frontend/{frontend_id}")
async def update_frontend_override(
    frontend_id: str,
    req: FrontendOverrideRequest,
    _: dict = Depends(require_admin),
):
    store = load_authorized_contacts()
    store.setdefault("per_frontend", {})[frontend_id] = {
        "mode": req.mode,
        "contacts": [c.model_dump() for c in req.contacts],
    }
    clean = save_authorized_contacts(store)
    return {"frontend_id": frontend_id, "override": clean["per_frontend"].get(frontend_id)}


@router.delete("/frontend/{frontend_id}")
async def delete_frontend_override(frontend_id: str, _: dict = Depends(require_admin)):
    store = load_authorized_contacts()
    pf = store.get("per_frontend") or {}
    if frontend_id in pf:
        del pf[frontend_id]
        store["per_frontend"] = pf
        save_authorized_contacts(store)
        logger.info(f"Contacts override removed for frontend {frontend_id}")
    return {"frontend_id": frontend_id, "removed": True}


@router.post("/frontend/{frontend_id}/copy-from/{src_frontend_id}")
async def copy_contacts_from_frontend(
    frontend_id: str,
    src_frontend_id: str,
    mode: Literal["replace", "append"] = Query("replace"),
    _: dict = Depends(require_admin),
):
    """Copy the contacts list from src_frontend_id into frontend_id."""
    store = load_authorized_contacts()
    src = (store.get("per_frontend") or {}).get(src_frontend_id)
    if not src:
        raise HTTPException(404, f"Source frontend '{src_frontend_id}' has no contacts override")
    store.setdefault("per_frontend", {})[frontend_id] = {
        "mode": mode,
        "contacts": list(src.get("contacts", [])),
    }
    clean = save_authorized_contacts(store)
    return {"frontend_id": frontend_id, "override": clean["per_frontend"].get(frontend_id)}


# --- Export ---

def _contacts_for_scope(store: dict[str, Any], scope: str) -> list[dict[str, str]]:
    if scope == "global":
        return store.get("global", [])
    if scope.startswith("frontend:"):
        fid = scope.split(":", 1)[1]
        return (store.get("per_frontend") or {}).get(fid, {}).get("contacts", [])
    raise HTTPException(400, f"Invalid scope: {scope}")


@router.get("/export")
async def export_contacts(
    scope: str = Query("global", description="global | frontend:{fid} | all"),
    _: dict = Depends(require_admin),
):
    """Export contacts as .xlsx. scope=all writes one sheet per scope."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    store = load_authorized_contacts()
    wb = Workbook()
    ws = wb.active
    assert ws is not None

    def _write_sheet(sheet, contacts: list[dict[str, str]]):
        sheet.append(list(_CONTACT_FIELDS))
        for c in contacts:
            sheet.append([c.get(f, "") for f in _CONTACT_FIELDS])

    if scope == "all":
        ws.title = "global"
        _write_sheet(ws, store.get("global", []))
        for fid, override in (store.get("per_frontend") or {}).items():
            # Excel sheet names limited to 31 chars, no special chars
            sheet_name = f"frontend_{fid}"[:31].replace(":", "_").replace("/", "_")
            s = wb.create_sheet(sheet_name)
            _write_sheet(s, override.get("contacts", []))
    else:
        contacts = _contacts_for_scope(store, scope)
        ws.title = scope.replace(":", "_")[:31]
        _write_sheet(ws, contacts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"authorized_contacts_{scope.replace(':', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Import (additive merge — never destructive) ---

def _parse_xlsx(data: bytes) -> tuple[list[dict[str, str]], int]:
    """Return (parsed_contacts, ignored_malformed). Uses first sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    if ws is None:
        return [], 0
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], 0
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    contacts: list[dict[str, str]] = []
    ignored = 0
    for row in rows[1:]:
        raw = {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}
        c = _normalise_contact(raw)
        if c:
            contacts.append(c)
        elif any(v not in (None, "") for v in row):
            ignored += 1
    return contacts, ignored


def _parse_csv(data: bytes) -> tuple[list[dict[str, str]], int]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    contacts: list[dict[str, str]] = []
    ignored = 0
    for raw in reader:
        norm_raw = {(k or "").strip().lower(): v for k, v in raw.items()}
        c = _normalise_contact(norm_raw)
        if c:
            contacts.append(c)
        elif any((v or "").strip() for v in raw.values()):
            ignored += 1
    return contacts, ignored


@router.post("/import")
async def import_contacts(
    file: UploadFile = File(...),
    scope: str = Query("global"),
    _: dict = Depends(require_admin),
):
    """Additive merge import from .xlsx or .csv.

    - Existing emails: fields updated where the incoming value is non-empty.
    - New emails: added.
    - Emails in backend but NOT in file: preserved (never deleted).
    """
    data = await file.read()
    name = (file.filename or "").lower()
    if name.endswith(".xlsx"):
        incoming, ignored = _parse_xlsx(data)
    elif name.endswith(".csv"):
        incoming, ignored = _parse_csv(data)
    else:
        raise HTTPException(400, "File must be .xlsx or .csv")

    store = load_authorized_contacts()

    if scope == "global":
        target_list = store.setdefault("global", [])
    elif scope.startswith("frontend:"):
        fid = scope.split(":", 1)[1]
        pf = store.setdefault("per_frontend", {})
        override = pf.setdefault(fid, {"mode": "replace", "contacts": []})
        target_list = override.setdefault("contacts", [])
    else:
        raise HTTPException(400, f"Invalid scope: {scope}")

    by_email: dict[str, dict[str, str]] = {c["email"]: c for c in target_list}
    added = 0
    updated = 0
    for new_c in incoming:
        existing = by_email.get(new_c["email"])
        if existing is None:
            by_email[new_c["email"]] = new_c
            added += 1
        else:
            changed = False
            for field in _CONTACT_FIELDS:
                if field == "email":
                    continue
                if new_c.get(field) and new_c[field] != existing.get(field, ""):
                    existing[field] = new_c[field]
                    changed = True
            if changed:
                updated += 1

    # Rebuild target_list preserving insertion of new entries after existing
    rebuilt = list(by_email.values())
    if scope == "global":
        store["global"] = rebuilt
    else:
        fid = scope.split(":", 1)[1]
        store["per_frontend"][fid]["contacts"] = rebuilt

    save_authorized_contacts(store)
    logger.info(f"Contacts import ({scope}): added={added} updated={updated} ignored={ignored}")
    return {"added": added, "updated": updated, "ignored_malformed": ignored, "scope": scope}
